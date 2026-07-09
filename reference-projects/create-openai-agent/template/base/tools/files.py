"""
Built-in file reading tools powered by markitdown.
Supports: PDF, Word (.docx), Excel (.xlsx), CSV, PowerPoint, images, HTML, ZIP.

Install: uv add 'openai-agents[file]'  or  uv add markitdown
"""

from __future__ import annotations
from pathlib import Path
from agents import function_tool


def _markitdown():
    try:
        from markitdown import MarkItDown
        return MarkItDown()
    except ImportError:
        raise ImportError(
            "markitdown not installed. Run: uv add 'create-agent[file]'  or  uv add markitdown"
        )


@function_tool
def read_file(file_path: str) -> str:
    """
    Read any file and return its content as markdown text.
    Supports: PDF, Word (.docx), Excel (.xlsx/.xls), CSV, PowerPoint (.pptx),
    images (with OCR), HTML, ZIP archives, and plain text files.
    """
    path = Path(file_path).expanduser()
    if not path.exists():
        return f"Error: file not found: {file_path}"

    # Plain text — no need for markitdown
    text_suffixes = {".txt", ".md", ".py", ".js", ".ts", ".json", ".yaml", ".yml", ".toml"}
    if path.suffix.lower() in text_suffixes:
        return path.read_text(encoding="utf-8", errors="replace")

    try:
        md = _markitdown()
        result = md.convert(str(path))
        return result.text_content or "(empty document)"
    except Exception as e:
        return f"Error reading {file_path}: {e}"


@function_tool
def read_pdf(file_path: str) -> str:
    """Read a PDF file and return its full text content as markdown."""
    return read_file.fn(file_path)


@function_tool
def read_csv(file_path: str, max_rows: int = 100) -> str:
    """
    Read a CSV file and return its content as a markdown table.
    max_rows: maximum number of rows to return (default 100).
    """
    path = Path(file_path).expanduser()
    if not path.exists():
        return f"Error: file not found: {file_path}"
    try:
        import csv
        with open(path, newline="", encoding="utf-8-sig") as f:
            reader = csv.reader(f)
            rows = list(reader)
        if not rows:
            return "(empty CSV)"
        header = rows[0]
        data = rows[1 : max_rows + 1]
        total = len(rows) - 1

        lines = ["| " + " | ".join(header) + " |"]
        lines.append("| " + " | ".join(["---"] * len(header)) + " |")
        for row in data:
            lines.append("| " + " | ".join(str(c) for c in row) + " |")
        if total > max_rows:
            lines.append(f"\n_Showing {max_rows} of {total} rows._")
        return "\n".join(lines)
    except Exception as e:
        return f"Error reading CSV: {e}"


@function_tool
def read_excel(file_path: str, sheet_name: str = "", max_rows: int = 100) -> str:
    """
    Read an Excel file (.xlsx/.xls) and return content as markdown tables.
    sheet_name: specific sheet to read (default: first sheet).
    max_rows: max rows per sheet (default 100).
    """
    path = Path(file_path).expanduser()
    if not path.exists():
        return f"Error: file not found: {file_path}"
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        sheets = [sheet_name] if sheet_name else wb.sheetnames
        output = []
        for name in sheets:
            ws = wb[name]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                output.append(f"## Sheet: {name}\n_(empty)_")
                continue
            header = [str(c) if c is not None else "" for c in rows[0]]
            data = rows[1 : max_rows + 1]
            lines = [f"## Sheet: {name}"]
            lines.append("| " + " | ".join(header) + " |")
            lines.append("| " + " | ".join(["---"] * len(header)) + " |")
            for row in data:
                lines.append("| " + " | ".join(str(c) if c is not None else "" for c in row) + " |")
            total = len(rows) - 1
            if total > max_rows:
                lines.append(f"\n_Showing {max_rows} of {total} rows._")
            output.append("\n".join(lines))
        return "\n\n".join(output)
    except ImportError:
        # Fallback to markitdown
        return read_file.fn(file_path)
    except Exception as e:
        return f"Error reading Excel: {e}"


@function_tool
def list_directory(dir_path: str, pattern: str = "*") -> str:
    """
    List files in a directory matching a glob pattern.
    Examples: list_directory('/data', '*.pdf') or list_directory('.', '*.csv')
    """
    path = Path(dir_path).expanduser()
    if not path.exists():
        return f"Error: directory not found: {dir_path}"
    files = sorted(path.glob(pattern))
    if not files:
        return f"No files matching '{pattern}' in {dir_path}"
    lines = [f"- {f.name} ({f.stat().st_size // 1024}KB)" for f in files if f.is_file()]
    return f"**{dir_path}** ({len(lines)} files):\n" + "\n".join(lines)


# Convenience bundle — import this in your agent
FILE_TOOLS = [read_file, read_pdf, read_csv, read_excel, list_directory]
